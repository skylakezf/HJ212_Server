import socket
import time
from openpyxl import load_workbook
from tqdm import tqdm  

# 使用前需要安装python3 
# 通过此命令安装程序依赖
#pip install openpyxl tqdm
#配置完成后通过此命令运行脚本 python main.py

def send_tcp_data(host, port, data):
    """
    通过TCP发送数据到指定的服务器

    :param host: 服务器地址
    :param port: 服务器端口
    :param data: 要发送的数据
    """
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as client_socket:
            client_socket.connect((host, port))
            client_socket.sendall(data.encode('utf-8'))
            print("\n")
            tqdm.write(f"已发送数据: {data}")
    except Exception as e:
        tqdm.write(f"发送数据时发生错误: {e}")

def main():
    # 配置
    excel_file = "data.xlsx"  # Excel 文件路径
    server_host = "113.45.10.40"  # 目标服务器地址
    server_port = 3000         # 目标服务器端口
    interval = 0.1            # 发送间隔（秒）
    # 此间隔可为空但为了防止由于网络阻塞建议此间隔填写为0.01

    try:
        # 加载 Excel 文件
        wb = load_workbook(excel_file)
        sheet = wb.active

        # 获取数据总行数
        max_row = sheet.max_row
        tqdm.write(f"发现 {max_row} 行数据，将每隔 {interval} 秒发送一次。")
        
        # 预估时间
        total_time = max_row * interval
        tqdm.write(f"预计完成时间: {total_time:.2f} 秒")
        print("按下CTRL+C取消运行")
        time.sleep(10)
        # 创建进度条
        with tqdm(total=max_row, desc="发送进度", unit="条", position=0, leave=True) as pbar:
            # 遍历第一列数据（从 A1 开始）
            for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
                cell_value = row[0]  # 取出第一列的值
                if cell_value is None:
                    tqdm.write("数据为空，停止发送。")
                    break
                data = str(cell_value).strip()
                send_tcp_data(server_host, server_port, data)
                time.sleep(interval)  # 等待指定时间间隔
                pbar.update(1)  # 更新进度条

        tqdm.write("所有数据已发送完毕。")
    except FileNotFoundError:
        tqdm.write(f"未找到文件: {excel_file}")
    except Exception as e:
        tqdm.write(f"程序运行时发生错误: {e}")

if __name__ == "__main__":
    main()
